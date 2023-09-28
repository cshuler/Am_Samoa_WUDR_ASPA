import os
import glob
import re
import shutil
import pandas as pd
from openpyxl import load_workbook


class FileManager:
    @staticmethod
    def delete_directory(path):
        if not os.path.exists(path):
            print(f"The path {path} does not exist")
        else:
            try:
                shutil.rmtree(path)
                print(f"Directory {path} has been deleted successfully")
            except OSError as e:
                print(f"Error: {e.filename} - {e.strerror}.")

    @staticmethod
    def copy_files(source_dir, destination_dir):
        if not os.path.exists(destination_dir):
            os.makedirs(destination_dir)
        for root, dirs, files in os.walk(source_dir):
            for file in files:
                source_path = os.path.join(root, file)
                destination_path = os.path.join(destination_dir, file)
                shutil.copy2(source_path, destination_path)

    @staticmethod
    def delete_non_spreadsheet_files(directory):
        all_files = glob.glob(os.path.join(directory, "*"))
        spreadsheet_files = glob.glob(os.path.join(directory, "*.csv")) + \
                            glob.glob(os.path.join(directory, "*.xlsx")) + \
                            glob.glob(os.path.join(directory, "*.xls"))
        non_spreadsheet_files = set(all_files) - set(spreadsheet_files)
        for file in non_spreadsheet_files:
            os.remove(file)

    @staticmethod
    def replace_month_names_in_files(path):
        months = {
            'january': '01', 'jan': '01', 'feb': '02', 'february': '02', 'mar': '03', 'march': '03', 'apr': '04',
            'april': '04', 'may': '05', 'jun': '06', 'june': '06', 'jul': '07', 'july': '07',
            'aug': '08', 'august': '08', 'sep': '09', 'september': '09', 'oct': '10', 'october': '10',
            'nov': '11', 'november': '11', 'dec': '12', 'december': '12'
        }
        for filename in os.listdir(path):
            if os.path.isdir(os.path.join(path, filename)):
                continue
            base_filename, file_extension = os.path.splitext(filename)
            new_base_filename = base_filename.lower()
            for month_name, month_number in months.items():
                new_base_filename = new_base_filename.replace(month_name, month_number)
            new_base_filename = re.sub(r'\(\d+\)', '', new_base_filename)
            new_base_filename = ''.join(c for c in new_base_filename if c.isdigit())
            new_filename = new_base_filename + file_extension
            if new_filename.lower() != filename.lower():
                shutil.move(os.path.join(path, filename), os.path.join(path, new_filename))

    @staticmethod
    def delete_files_with_long_names(path):
        for filename in os.listdir(path):
            if os.path.isdir(os.path.join(path, filename)):
                continue
            file_name_without_extension = os.path.splitext(filename)[0]
            if len(file_name_without_extension) > 6:
                os.remove(os.path.join(path, filename))

    @staticmethod
    def rename_dates_in_files(path):
        regex_mm_yyyy = re.compile(r'^(01|02|03|04|05|06|07|08|09|10|11|12)(20\d{2})$')
        regex_yyyy_mm = re.compile(r'^(20\d{2})(01|02|03|04|05|06|07|08|09|10|11|12)$')
        for filename in os.listdir(path):
            if os.path.isdir(os.path.join(path, filename)):
                continue
            name_without_extension, extension = os.path.splitext(filename)
            match = regex_mm_yyyy.match(name_without_extension)
            if match:
                new_filename = f"{match.group(2)}-{match.group(1)}{extension}"
                os.rename(os.path.join(path, filename), os.path.join(path, new_filename))
                continue
            match = regex_yyyy_mm.match(name_without_extension)
            if match:
                new_filename = f"{match.group(1)}-{match.group(2)}{extension}"
                os.rename(os.path.join(path, filename), os.path.join(path, new_filename))

    @staticmethod
    def remove_all_but_second_sheet_in_xlsx(path):
        for filename in os.listdir(path):
            if filename.endswith(".xlsx"):
                file_path = os.path.join(path, filename)
                workbook = load_workbook(file_path)
                if len(workbook.sheetnames) < 2:
                    print(f"Skipping {filename} because it has less than two sheets")
                    continue
                second_sheet = workbook.sheetnames[1]
                for sheet in workbook.sheetnames:
                    if sheet != second_sheet:
                        del workbook[sheet]
                workbook.save(file_path)


class DataProcessor:
    @staticmethod
    def load_and_combine_files(directory_path):
        all_files = [f for f in os.listdir(directory_path) if
                     os.path.isfile(os.path.join(directory_path, f)) and f.endswith('.xlsx')]
        dataframes = []
        for file in all_files:
            df = pd.read_excel(os.path.join(directory_path, file))
            dataframes.append(df)
        combined_df = pd.concat(dataframes, ignore_index=True)
        return combined_df

    @staticmethod
    def export_to_csv(df, output_path):
        df.to_csv(output_path, index=False)
        print(f"Data exported to: {output_path}")


class DataCleaner:
    @staticmethod
    def load_data(file_path):
        return pd.read_csv(file_path, low_memory=False)

    @staticmethod
    def clean_data(cs_data):
        cs_data = cs_data[cs_data['LOSADD'].notnull()]
        cs_data = cs_data.rename(columns=lambda x: x.strip())
        cs_data['Usage'] = cs_data['Usage'].astype(str).str.replace(" ", "")
        cs_data['Usage'] = cs_data['Usage'].str.replace("-", "0")
        cs_data['Usage'] = cs_data['Usage'].str.replace(",", "")
        cs_data['Usage'] = pd.to_numeric(cs_data['Usage'])
        return cs_data

    @staticmethod
    def filter_for_residential(cs_data):
        return cs_data[cs_data['RSP'] == 'RES']

    @staticmethod
    def map_rsp_to_beneficial_use_category(cs_data):
        rsp_mapping = {
            'RES': 'Domestic',
            'IND': 'Industrial',
            'LGS': 'Industrial',
            'SGS': 'Commercial',
            'ASG': 'Commercial'
        }
        cs_data['BeneficialUseCategory'] = cs_data['RSP'].map(rsp_mapping)
        return cs_data

    @staticmethod
    def transform_to_site_specific_format(cs_data):
        # Define an empty dataframe with the site-specific columns
        all_columns = [
            "MethodUUID", "OrganizationUUID", "SiteUUID", "VariableSpecificUUID", "WaterSourceUUID", "Amount",
            "AllocationCropDutyAmount", "AssociatedNativeAllocationIDs", "BeneficialUseCategory",
            "CommunityWaterSupplySystem", "CropTypeCV", "CustomerTypeCV", "DataPublicationDate", "DataPublicationDOI",
            "Geometry", "IrrigatedAcreage", "IrrigationMethodCV", "PopulationServed", "PowerGeneratedGWh",
            "PowerType", "PrimaryUseCategory", "ReportYearCV", "SDWISIdentifier", "TimeframeEnd", "TimeframeStart"
        ]

        village_key = {
            'AASU': "Aasu",
            'AASU FOU': 'Aasu',
            'AFAO': 'Afao',
            'AFONO': 'Afono',
            'AGUGULU': 'Agugulu',
            'ALAO': 'Alao',
            'ALEGA': 'Alega',
            'ALOFAU': 'Alofau',
            'AMALUIA': 'Amaluia',
            'AMANAVE': 'Amanave',
            'AMAUA': 'Amaua',
            'AMOULI': 'Amouli',
            'AOA': 'Aoa',
            'AOLOAU': 'Aoloau',
            'ASILI': 'Asili',
            'ATAULOMA': 'Afao',
            'ATUU': 'Atuu',
            'AUA': 'Aua',
            'AUASI': 'Auasi',
            'AUNUU': 'Aunuu',
            'AUTO': 'Auto',
            'AVAIO': 'Avaio',
            'FAGAALU': 'Fagaalu',
            'FAGAITUA': 'Fagaitua',
            'FAGALII': 'Fagalii',
            'FAGAMALO': 'Fagamalo',
            'FAGANEANEA': 'Faganeanea',
            'FAGASA': 'Fagasa',
            'FAGATOGO': 'Fagatogo',
            'FAILOLO': 'Failolo',
            'FALEASAO': 'Tau',
            'FALENIU': 'Faleniu',
            'FATUMAFUTI': 'Fatumafuti',
            'FITIUTA': 'Tau',
            'FOGAGOGO': 'Iliili',
            'FUTIGA': 'Futiga',
            'Fagaalu': 'Fagaalu',
            'GATAIVAI': 'Utulei',
            'ILIILI': 'Iliili',
            'LAULII': 'Laulii',
            'LELOALOA': 'Leloaloa',
            'LEONE': 'Leone',
            'MALAEIMI': 'Malaeimi',
            'MALAELOA': 'Malaeloa',
            'MALALOA': 'Fagatogo',
            'MALOATA': 'Maloata',
            'MAPUSAGA': 'Mapusagafou',
            'MAPUSAGA FOU': 'Mapusagafou',
            'MASAUSI': 'Masausi',
            'MASEFAU': 'Masefau',
            'MATUU': 'Matuu',
            'MESEPA': 'Mesepa',
            'NUA': 'Nua',
            'NUUULI': 'Nuuuli',
            'OFU': 'Ofu',
            'OLOSEGA': 'Olosega',
            'ONENOA': 'Onenoa',
            'PAGAI': 'Pagai',
            'PAGO PAGO': 'Pago Pago',
            'PAVAIAI': 'Pavaiai',
            'POLOA': 'Poloa',
            'SAILELE': 'Sailele',
            'SATALA': 'Pago Pago',
            'SEETAGA': 'Seetaga',
            'TAFETA': 'Mapusagafou',
            'TAFUNA': 'Tafuna',
            'TAPUTIMU': 'Taputimu',
            'TAU': 'Tau',
            'TULA': 'Tula',
            'UTULEI': 'Utulei',
            'UTUMEA': 'Utumea West',
            'UTUMEA-SASAE': 'Utumea East',
            'UTUSIA': 'Fagaitua',
            'VAILOA': 'Vailoatai',
            'VAITOGI': 'Vaitogi',
            'VATIA': 'Vatia',
            'Vaitogi': 'Vaitogi'
        }
        transformed_data = pd.DataFrame(columns=all_columns)

        # Populate the columns based on the cs_data
        transformed_data['SiteUUID'] = cs_data['LOSADD'].map(village_key)
        transformed_data['VariableSpecificUUID'] = 'UTssps_V3'
        transformed_data['Amount'] = cs_data['Usage']
        transformed_data['BeneficialUseCategory'] = cs_data['BeneficialUseCategory']

        # Calculate TimeframeStart and TimeframeEnd
        transformed_data['TimeframeStart'] = pd.to_datetime(cs_data['Rdg Date']) - pd.to_timedelta(
            cs_data['Days of Usage'], unit='D')
        transformed_data['TimeframeEnd'] = cs_data['Rdg Date']

        # The rest of the columns will remain NaN
        return transformed_data


if __name__ == '__main__':
    # File management steps
    FileManager.delete_directory("OrganizedData")
    FileManager.copy_files("OriginalData\\CS Billing", "OrganizedData\\WaterBilling")
    FileManager.delete_non_spreadsheet_files("OrganizedData\\WaterBilling")
    FileManager.replace_month_names_in_files("OrganizedData\\WaterBilling")
    FileManager.delete_files_with_long_names("OrganizedData\\WaterBilling")
    FileManager.rename_dates_in_files("OrganizedData\\WaterBilling")

    # Load and export combined data
    combined_data = DataProcessor.load_and_combine_files("OrganizedData\\WaterBilling")
    DataProcessor.export_to_csv(combined_data, "OrganizedData\\data.csv")

    # Data cleaning and transformation steps
    cs_data = DataCleaner.load_data("OrganizedData\\data.csv")
    cs_data = DataCleaner.clean_data(cs_data)
    cs_data = DataCleaner.filter_for_residential(cs_data)
    cs_data = DataCleaner.map_rsp_to_beneficial_use_category(cs_data)
    cs_data = DataCleaner.transform_to_site_specific_format(cs_data)

    DataProcessor.export_to_csv(cs_data, "OrganizedData\\formatted_data.csv")
